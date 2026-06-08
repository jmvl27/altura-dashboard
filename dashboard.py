#!/usr/bin/env python3
"""
Altura Research — Dashboard Campaña (GitHub Actions version)
Descarga el Excel desde Google Drive, genera el HTML y lo sube por FTP.
"""

import ftplib
import io
import os
import requests
import openpyxl
from collections import Counter
from datetime import datetime

DRIVE_FILE_ID = "10VmTTlTzO9hNXf5ktIssplnh8qNBIARm"
HTML_FILE     = "Dashboard_Campana.html"

def download_excel():
    """Descarga el Excel desde Google Drive (link público)."""
    print("📥 Descargando Excel desde Google Drive...")
    url = f"https://drive.google.com/uc?export=download&id={DRIVE_FILE_ID}"
    session = requests.Session()
    response = session.get(url, stream=True)

    # Manejar confirmación de virus scan para archivos grandes
    for key, value in response.cookies.items():
        if key.startswith("download_warning"):
            url = f"https://drive.google.com/uc?export=download&confirm={value}&id={DRIVE_FILE_ID}"
            response = session.get(url, stream=True)
            break

    content = b"".join(response.iter_content(chunk_size=32768))
    print(f"✅ Excel descargado ({len(content):,} bytes)")
    return io.BytesIO(content)


def generate_dashboard(excel_bytes):
    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))

    contacts = [r for r in rows if r.get("ID") != 999]
    total = len(contacts)

    def cnt(col, val=1):
        return sum(1 for r in contacts if r.get(col) == val or r.get(col) == float(val))

    conocidos     = cnt("CONOCIDO")
    por_contactar = sum(1 for r in contacts if r.get("CONOCIDO") != 1 and not r.get("CONTACTADO"))
    contactados   = cnt("CONTACTADO")
    reuniones     = cnt("REUNIÓN AGENDADA")
    propuestas    = cnt("ENVÍO DE PROPUESTA")
    negociacion   = cnt("NEGOCIACIÓN")
    perdidos      = cnt("PERDIDO")
    ganados       = cnt("GANADO")
    decisores     = cnt("DECISOR")
    sin_correo    = sum(1 for r in contacts if not r.get("CORREO"))

    sectores = Counter(r.get("SECTOR") for r in contacts if r.get("SECTOR"))
    sectores_sorted = sorted(sectores.items(), key=lambda x: -x[1])

    tasa_conv    = round(ganados / contactados * 100, 1) if contactados else 0
    tasa_reunion = round(reuniones / contactados * 100, 1) if contactados else 0
    generated    = datetime.now().strftime("%d/%m/%Y %H:%M")

    funnel = [
        ("Base total",            total,         "#1B2F55"),
        ("Por contactar",         por_contactar, "#3B6CB5"),
        ("Contactados",           contactados,   "#2563EB"),
        ("Reunión agendada",      reuniones,     "#F59E0B"),
        ("Propuesta enviada",     propuestas,    "#8B5CF6"),
        ("Negociación",           negociacion,   "#F47B20"),
        ("Ganados",               ganados,       "#10B981"),
    ]
    max_funnel = max(v for _, v, _ in funnel) or 1

    funnel_bars = ""
    for label, val, color in funnel:
        pct = round(val / max_funnel * 100)
        pct_total = round(val / total * 100, 1) if total else 0
        funnel_bars += f"""
      <div class="funnel-row">
        <div class="funnel-label">{label}</div>
        <div class="funnel-bar-wrap">
          <div class="funnel-bar" style="width:{pct}%;background:{color}"></div>
        </div>
        <div class="funnel-val">{val:,} <span class="funnel-pct">({pct_total}%)</span></div>
      </div>"""

    sector_rows = "".join(
        f"<tr><td>{s}</td><td>{n:,}</td><td>{round(n/total*100,1)}%</td></tr>"
        for s, n in sectores_sorted[:10]
    )

    alert = "" if contactados > 0 else \
        '<div class="alert">⚠️ Campaña aún no iniciada — ningún contacto enviado todavía.</div>'

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Campaña Email — Altura Research</title>
<style>
  *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
  :root{{--navy:#1B2F55;--blue:#2563EB;--light:#F0F4FB;--gray:#64748B;--text:#1E293B;--white:#fff;--border:#E2E8F0;--green:#10B981;--orange:#F47B20}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#F3F4F6;color:var(--text);line-height:1.6}}
  header{{background:var(--navy);color:#fff;padding:20px 32px;display:flex;justify-content:space-between;align-items:center}}
  header h1{{font-size:1.3rem;font-weight:700}}
  header .sub{{font-size:.78rem;opacity:.65;margin-top:2px}}
  .container{{max-width:1100px;margin:0 auto;padding:24px}}
  .kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px}}
  .kpi{{background:#fff;border-radius:10px;padding:18px 20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .kpi .val{{font-size:2rem;font-weight:700;color:var(--navy);line-height:1}}
  .kpi .lbl{{font-size:.72rem;color:var(--gray);text-transform:uppercase;letter-spacing:.05em;margin-top:6px}}
  .kpi.blue .val{{color:var(--blue)}}
  .kpi.orange .val{{color:var(--orange)}}
  .grid2{{display:grid;grid-template-columns:1.6fr 1fr;gap:20px;margin-bottom:22px}}
  .card{{background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .card h2{{font-size:.82rem;font-weight:700;color:var(--gray);text-transform:uppercase;letter-spacing:.06em;margin-bottom:16px}}
  .funnel-row{{display:flex;align-items:center;gap:10px;margin-bottom:10px}}
  .funnel-label{{width:180px;font-size:.82rem;color:var(--text);flex-shrink:0}}
  .funnel-bar-wrap{{flex:1;background:#F3F4F6;border-radius:4px;height:16px;overflow:hidden}}
  .funnel-bar{{height:100%;border-radius:4px;min-width:3px}}
  .funnel-val{{width:90px;text-align:right;font-size:.82rem;font-weight:600}}
  .funnel-pct{{font-weight:400;color:var(--gray)}}
  table{{width:100%;border-collapse:collapse;font-size:.85rem}}
  th{{background:#F9FAFB;padding:8px 12px;text-align:left;font-weight:600;color:var(--gray);font-size:.72rem;text-transform:uppercase;border-bottom:2px solid var(--border)}}
  td{{padding:8px 12px;border-bottom:1px solid #F3F4F6}}
  .stat-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}
  .stat{{background:#F9FAFB;border-radius:8px;padding:14px;text-align:center}}
  .stat .n{{font-size:1.5rem;font-weight:700;color:var(--navy)}}
  .stat .l{{font-size:.75rem;color:var(--gray);margin-top:3px}}
  .alert{{background:#FEF3C7;border:1px solid #FCD34D;border-radius:8px;padding:12px 16px;font-size:.83rem;color:#92400E;margin-bottom:22px}}
  footer{{text-align:center;font-size:.72rem;color:#9CA3AF;padding:20px}}
  @media(max-width:700px){{.kpis,.stat-grid{{grid-template-columns:1fr 1fr}}.grid2{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<header>
  <div>
    <h1>Dashboard Campaña Email — Altura Research</h1>
    <div class="sub">Actualizado: {generated}</div>
  </div>
</header>
<div class="container">
  {alert}
  <div class="kpis">
    <div class="kpi"><div class="val">{total:,}</div><div class="lbl">Total contactos</div></div>
    <div class="kpi blue"><div class="val">{por_contactar:,}</div><div class="lbl">Por contactar</div></div>
    <div class="kpi"><div class="val">{decisores:,}</div><div class="lbl">Decisores</div></div>
    <div class="kpi orange"><div class="val">{sin_correo}</div><div class="lbl">Sin correo</div></div>
  </div>
  <div class="grid2">
    <div class="card">
      <h2>Funnel de la campaña</h2>
      {funnel_bars}
    </div>
    <div class="card">
      <h2>Por sector</h2>
      <table>
        <thead><tr><th>Sector</th><th>N</th><th>%</th></tr></thead>
        <tbody>{sector_rows}</tbody>
      </table>
    </div>
  </div>
  <div class="card" style="margin-bottom:22px">
    <h2>Métricas de conversión</h2>
    <div class="stat-grid">
      <div class="stat"><div class="n">{contactados:,}</div><div class="l">Contactados</div></div>
      <div class="stat"><div class="n">{reuniones}</div><div class="l">Reuniones agendadas</div></div>
      <div class="stat"><div class="n">{propuestas}</div><div class="l">Propuestas enviadas</div></div>
      <div class="stat"><div class="n">{negociacion}</div><div class="l">En negociación</div></div>
      <div class="stat"><div class="n" style="color:var(--green)">{ganados}</div><div class="l">Ganados</div></div>
      <div class="stat"><div class="n" style="color:#EF4444">{perdidos}</div><div class="l">Perdidos</div></div>
      <div class="stat"><div class="n">{tasa_reunion}%</div><div class="l">Contacto → reunión</div></div>
      <div class="stat"><div class="n">{tasa_conv}%</div><div class="l">Tasa de cierre</div></div>
    </div>
  </div>
</div>
<footer>Altura Research · Dashboard generado automáticamente</footer>
</body>
</html>"""

    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Dashboard generado: {HTML_FILE}")


def upload_ftp():
    host     = os.environ["FTP_HOST"]
    user     = os.environ["FTP_USER"]
    password = os.environ["FTP_PASS"]
    path     = "/dashboard_ventas"

    print(f"📡 Conectando a {host}...")
    ftp = ftplib.FTP()
    ftp.connect(host, 21, timeout=30)
    ftp.login(user, password)
    ftp.set_pasv(True)
    print(f"✅ Conectado como {user}")
    try:
        ftp.mkd(path)
    except ftplib.error_perm:
        pass
    ftp.cwd(path)
    with open(HTML_FILE, "rb") as f:
        ftp.storbinary("STOR Dashboard_Campana.html", f)
    ftp.quit()
    print(f"✅ Subido a {host}{path}/Dashboard_Campana.html")


if __name__ == "__main__":
    import sys
    no_ftp = "--no-ftp" in sys.argv
    print(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} — Iniciando...")
    sys.stdout.flush()
    try:
        excel_bytes = download_excel()
        sys.stdout.flush()
        generate_dashboard(excel_bytes)
        sys.stdout.flush()
        if not no_ftp:
            upload_ftp()
            sys.stdout.flush()
        print("✅ Listo.")
    except Exception as e:
        import traceback
        print(f"❌ ERROR: {e}")
        traceback.print_exc()
        sys.exit(1)
